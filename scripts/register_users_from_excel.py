from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Reuse the project's Bitrix REST client. It already supports retries and the
# Windows corporate certificate store through truststore.
from src.bitrix import BitrixClient


SHEET_NAME = "Пользователи"
OUTPUT_DIR = Path("output/user-registration")


def normalize(value: Any) -> str:
    text = str(value or "").strip().casefold().replace("ё", "е")
    return " ".join(text.split())


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_email(value: Any) -> str:
    return clean(value).casefold()


def valid_email(value: str) -> bool:
    return bool(re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", value))


HEADER_ALIASES = {
    "last_name": {"фамилия", "last_name", "lastname"},
    "name": {"имя", "name", "first_name", "firstname"},
    "second_name": {"отчество", "second_name", "middlename", "middle_name"},
    "email": {"email", "e-mail", "почта", "электронная почта"},
    "department": {"подразделение", "департамент", "отдел", "department"},
    "department_id": {
        "id подразделения",
        "id департамента",
        "department_id",
        "department id",
    },
}


@dataclass
class InputUser:
    row_number: int
    last_name: str
    name: str
    second_name: str
    email: str
    department: str
    department_id: str

    @property
    def full_name(self) -> str:
        return " ".join(x for x in (self.last_name, self.name, self.second_name) if x)


@dataclass
class Action:
    row_number: int
    full_name: str
    email: str
    department: str
    department_id: str
    status: str
    target_user_id: str = ""
    message: str = ""


def resolve_headers(values: list[Any]) -> dict[str, int]:
    normalized = {normalize(value): index for index, value in enumerate(values) if clean(value)}
    result: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        matches = [normalized[normalize(alias)] for alias in aliases if normalize(alias) in normalized]
        if matches:
            result[field] = matches[0]
    missing = [field for field in ("last_name", "name", "email", "department") if field not in result]
    if missing:
        raise ValueError(
            "В листе 'Пользователи' отсутствуют обязательные колонки: " + ", ".join(missing)
        )
    return result


def read_users(path: Path) -> list[InputUser]:
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"В файле должен быть лист '{SHEET_NAME}'")
    sheet = workbook[SHEET_NAME]
    header_values = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    columns = resolve_headers(header_values)

    users: list[InputUser] = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        def get(field: str) -> str:
            index = columns.get(field)
            return clean(cells[index]) if index is not None and index < len(cells) else ""

        values = {
            "last_name": get("last_name"),
            "name": get("name"),
            "second_name": get("second_name"),
            "email": get("email"),
            "department": get("department"),
            "department_id": get("department_id"),
        }
        if not any(values.values()):
            continue
        users.append(InputUser(row_number=row_number, **values))
    workbook.close()
    return users


def user_full_name(user: dict[str, Any]) -> str:
    return " ".join(
        x for x in (clean(user.get("LAST_NAME")), clean(user.get("NAME")), clean(user.get("SECOND_NAME"))) if x
    )


def write_reports(actions: list[Action], mode: str, source_file: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    summary = {
        "mode": mode,
        "source_file": source_file,
        "total_rows": len(actions),
        "counts_by_status": {},
    }
    for action in actions:
        summary["counts_by_status"][action.status] = summary["counts_by_status"].get(action.status, 0) + 1

    (OUTPUT_DIR / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    fields = list(Action.__dataclass_fields__.keys())
    with (OUTPUT_DIR / "actions.csv").open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(asdict(action) for action in actions)

    errors = [action for action in actions if action.status == "ERROR"]
    with (OUTPUT_DIR / "errors.csv").open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(asdict(action) for action in errors)


def main() -> int:
    parser = argparse.ArgumentParser(description="Регистрация пользователей Bitrix24 из Excel")
    parser.add_argument("--file", default="data/users_to_invite.xlsx", help="Путь к Excel-файлу")
    parser.add_argument("--mode", choices=("dry_run", "apply"), default="dry_run")
    args = parser.parse_args()

    source_path = Path(args.file)
    actions: list[Action] = []

    try:
        input_users = read_users(source_path)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    if not input_users:
        print("ERROR: В Excel нет пользователей для обработки.", file=sys.stderr)
        return 2

    client = BitrixClient.from_env()
    departments = client.list_all("department.get", {})
    target_users = client.list_all("user.get", {})

    departments_by_id = {clean(item.get("ID")): item for item in departments if clean(item.get("ID"))}
    departments_by_name: dict[str, list[dict[str, Any]]] = {}
    for department in departments:
        departments_by_name.setdefault(normalize(department.get("NAME")), []).append(department)

    users_by_email = {
        normalize_email(item.get("EMAIL")): item
        for item in target_users
        if normalize_email(item.get("EMAIL"))
    }
    users_by_name: dict[str, list[dict[str, Any]]] = {}
    for item in target_users:
        full_name = normalize(user_full_name(item))
        if full_name:
            users_by_name.setdefault(full_name, []).append(item)

    seen_emails: set[str] = set()

    for row in input_users:
        email = normalize_email(row.email)
        action_base = {
            "row_number": row.row_number,
            "full_name": row.full_name,
            "email": row.email,
            "department": row.department,
            "department_id": row.department_id,
        }

        if not row.last_name or not row.name:
            actions.append(Action(**action_base, status="ERROR", message="Не заполнены фамилия или имя"))
            continue
        if not valid_email(email):
            actions.append(Action(**action_base, status="ERROR", message="Некорректный email"))
            continue
        if email in seen_emails:
            actions.append(Action(**action_base, status="ERROR", message="Повтор email внутри Excel"))
            continue
        seen_emails.add(email)

        existing = users_by_email.get(email)
        if existing:
            actions.append(
                Action(
                    **action_base,
                    status="SKIP",
                    target_user_id=clean(existing.get("ID")),
                    message="Пользователь с таким email уже существует",
                )
            )
            continue

        same_name = users_by_name.get(normalize(row.full_name), [])
        if same_name:
            existing_emails = ", ".join(clean(item.get("EMAIL")) for item in same_name)
            actions.append(
                Action(
                    **action_base,
                    status="ERROR",
                    message=f"В Bitrix уже найдено такое ФИО с другим email: {existing_emails}",
                )
            )
            continue

        department: dict[str, Any] | None = None
        if row.department_id:
            department = departments_by_id.get(row.department_id)
            if department is None:
                actions.append(
                    Action(**action_base, status="ERROR", message="ID подразделения не найден в Bitrix24")
                )
                continue
        else:
            matches = departments_by_name.get(normalize(row.department), [])
            if len(matches) == 1:
                department = matches[0]
            elif not matches:
                actions.append(
                    Action(**action_base, status="ERROR", message="Подразделение не найдено по точному названию")
                )
                continue
            else:
                ids = ", ".join(clean(item.get("ID")) for item in matches)
                actions.append(
                    Action(
                        **action_base,
                        status="ERROR",
                        message=f"Найдено несколько подразделений с таким названием. Укажите ID: {ids}",
                    )
                )
                continue

        resolved_department_id = clean(department.get("ID"))
        resolved_department_name = clean(department.get("NAME"))
        action_base["department"] = resolved_department_name
        action_base["department_id"] = resolved_department_id

        fields: dict[str, Any] = {
            "EMAIL": row.email.strip(),
            "NAME": row.name.strip(),
            "LAST_NAME": row.last_name.strip(),
            "UF_DEPARTMENT": [int(resolved_department_id)],
        }
        if row.second_name:
            fields["SECOND_NAME"] = row.second_name.strip()

        if args.mode == "dry_run":
            actions.append(
                Action(**action_base, status="DRY_RUN", message="Готов к регистрации и отправке приглашения")
            )
            continue

        try:
            result = client.call("user.add", fields)
            target_id = clean(result)
            actions.append(
                Action(
                    **action_base,
                    status="OK",
                    target_user_id=target_id,
                    message="Пользователь создан, приглашение отправлено",
                )
            )
        except Exception as exc:
            actions.append(Action(**action_base, status="ERROR", message=str(exc)))

    write_reports(actions, args.mode, str(source_path))

    for action in actions:
        print(
            f"{action.status:7} | row {action.row_number:>3} | {action.email:<40} | "
            f"{action.department} ({action.department_id}) | {action.message}"
        )

    errors = sum(1 for action in actions if action.status == "ERROR")
    print(f"\nProcessed: {len(actions)}; errors: {errors}; mode: {args.mode}")
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
